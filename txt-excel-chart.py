import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart3D, Reference

def read_txt_and_convert_to_csv(input_file, output_file):
    """
    Reads a text file and converts it to CSV format by adding commas between each word.
    
    Parameters:
    input_file (str): Path to the input text file.
    output_file (str): Path to the output CSV file.
    """
    with open(input_file, 'r') as txt_file, open(output_file, 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        
        for line in txt_file:
            words = line.strip().split()
            csv_writer.writerow(words)

def create_excel_with_3d_clustered_chart(csv_file, excel_file):
    """
    Reads a CSV file and writes it to an Excel file, adding a 3D clustered column chart to it.
    
    Parameters:
    csv_file (str): Path to the input CSV file.
    excel_file (str): Path to the output Excel file.
    """
    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(csv_file)
    
    # Write the DataFrame to an Excel file
    df.to_excel(excel_file, index=False, engine='openpyxl')
    
    # Load the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Create a 3D clustered column chart
    chart = BarChart3D()
    chart.type = "col"
    chart.grouping = "clustered"
    
    # Specify the data for the chart
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    
    # Specify the categories (labels) for the chart
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.set_categories(categories)
    
    # Add the chart to the worksheet
    ws.add_chart(chart, "E5")  # Position the chart at cell E5
    
    # Save the Excel file
    wb.save(excel_file)

# Example usage:
input_file = 'C:/Users/prana/data.txt'  # Update this path
output_file = 'C:/Users/prana/output.csv'  # Update this path
excel_file = 'C:/Users/prana/output.xlsx'  # Update this path

# Convert the text file to CSV
read_txt_and_convert_to_csv(input_file, output_file)

# Create an Excel file with a 3D clustered column chart
create_excel_with_3d_clustered_chart(output_file, excel_file)
